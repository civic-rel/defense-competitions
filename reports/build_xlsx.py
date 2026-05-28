"""Multi-sheet XLSX export.

Companion to the PDF report. Carries the tabular content that was
crowding the PDF — master list, cross-event participants, raw
evidence — plus a focused Outreach sheet built for paste-into-CRM
custom outbound. Five sheets, one logical view each:

    Outreach        — focused custom-outreach view:
                      Company / Last seen / Event / Website / contact
    Master list     — every participating company (former PDF
                      Section 2, minus the Type column)
    Cross-event     — companies appearing at 2+ events (former
                      PDF Section 3)
    Emerging        — emerging / unverified participants (Section 4)
    Raw evidence    — every Participation row, flattened

Run:
    python -m reports.build_xlsx
    # writes reports/out/report_<YYYY-MM-DD>.xlsx
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.build_markdown import (
    _CONF_RANK,
    _filter_data_for_target_audience,
    _highest_confidence,
    _highest_role,
    _load_all,
)


# ---- Styling helpers ----

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="333333")
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_BODY_ALIGN = Alignment(vertical="top", wrap_text=True)


def _write_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[list],
    widths: list[int] | None = None,
) -> None:
    """Add a sheet with header styling + reasonable column widths."""
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    for r in rows:
        ws.append([
            v if v not in (None, "") else "" for v in r
        ])
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col_idx).alignment = _BODY_ALIGN
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ---- Helpers to surface contact-route + last-event ----

def _last_event(data: dict, ps: list[dict]) -> tuple[str, str]:
    """Return (last_seen_date_iso, last_event_name) for a company,
    based on the events its participations attach to. Picks the
    event with the latest dates_start.
    """
    best_date = ""
    best_name = ""
    for p in ps:
        ev = data["event_by_id"].get(p["event_id"])
        if not ev:
            continue
        ds = ev.get("dates_start") or ""
        if ds > best_date:
            best_date = ds
            best_name = ev.get("name", "")
    return best_date, best_name


def _website_or_github(company: dict) -> str:
    """Same priority as the PDF's _website_display(), inlined here
    to avoid a circular import with build_pdf."""
    website = (company.get("website") or "").strip()
    if website:
        return website
    aliases = company.get("aliases") or []
    if isinstance(aliases, str):
        try:
            import json
            aliases = json.loads(aliases)
        except (ValueError, TypeError):
            aliases = []
    for a in aliases:
        if isinstance(a, str) and "github.com/" in a:
            return a
    return ""


# ---- Sheet builders ----

def _sheet_outreach(wb: Workbook, data: dict) -> int:
    """Focused outreach sheet: one row per company, with the
    most recent event for personalization."""
    rows = []
    for c in sorted(
        data["companies"],
        key=lambda x: -len(data["by_company"][x["id"]]),
    ):
        ps = data["by_company"][c["id"]]
        if not ps:
            continue
        last_date, last_event_name = _last_event(data, ps)
        rows.append([
            c["name"],
            last_date,
            last_event_name,
            _website_or_github(company=c),
        ])
    _write_sheet(
        wb, "Outreach",
        ["Company", "Last seen (date)", "Last event", "Website / contact"],
        rows,
        widths=[28, 14, 50, 50],
    )
    return len(rows)


def _sheet_master_list(wb: Workbook, data: dict) -> int:
    """Former PDF Section 2 — every participating company minus
    the Type column (user feedback: not beneficial).
    """
    rows = []
    for c in sorted(
        data["companies"],
        key=lambda x: -len(data["by_company"][x["id"]]),
    ):
        ps = data["by_company"][c["id"]]
        if not ps:
            continue
        event_ids = {p["event_id"] for p in ps}
        where_seen = "; ".join(sorted(
            data["event_by_id"].get(eid, {}).get("name", eid)
            for eid in event_ids
        ))
        rows.append([
            c["name"],
            ", ".join(c.get("domains") or []),
            len(event_ids),
            where_seen,
            _highest_role(ps),
            _highest_confidence(ps),
            ", ".join(c.get("notable_investors") or []),
            c.get("website") or "",
            c.get("crunchbase_url") or "",
        ])
    _write_sheet(
        wb, "Master list",
        [
            "Company", "Domains", "Events", "Where seen",
            "Top role", "Confidence", "Investors",
            "Website", "Crunchbase URL",
        ],
        rows,
        widths=[28, 22, 8, 60, 12, 16, 30, 35, 35],
    )
    return len(rows)


def _sheet_cross_event(wb: Workbook, data: dict) -> int:
    """Former PDF Section 3 — companies appearing at 2+ events."""
    rows = []
    for c in data["companies"]:
        events = {p["event_id"] for p in data["by_company"][c["id"]]}
        if len(events) < 2:
            continue
        names = sorted(
            data["event_by_id"].get(eid, {}).get("name", eid)
            for eid in events
        )
        rows.append([c["name"], len(events), "; ".join(names)])
    rows.sort(key=lambda r: -int(r[1]))
    _write_sheet(
        wb, "Cross-event",
        ["Company", "Event count", "Events"],
        rows,
        widths=[28, 12, 80],
    )
    return len(rows)


def _sheet_emerging(wb: Workbook, data: dict) -> int:
    """Former PDF Section 4 — emerging / unverified participants
    using the same definition as the PDF.
    """
    from reports.build_pdf import _is_emerging_unverified, _emerging_github_url
    rows = []
    for c in data["companies"]:
        if not _is_emerging_unverified(c):
            continue
        ps = data["by_company"][c["id"]]
        events = {p["event_id"] for p in ps}
        rows.append([
            c["name"],
            len(events),
            _highest_confidence(ps),
            _emerging_github_url(c),
            ", ".join(
                data["event_by_id"].get(eid, {}).get("name", eid)
                for eid in events
            ),
        ])
    _write_sheet(
        wb, "Emerging",
        ["Project / team", "Events", "Confidence", "GitHub", "Where seen"],
        rows,
        widths=[26, 8, 16, 40, 60],
    )
    return len(rows)


def _sheet_raw_evidence(wb: Workbook, data: dict) -> int:
    """Flattened raw evidence — one row per Participation."""
    rows = []
    by_company = data["by_company"]
    name_of = lambda cid: data["company_by_id"].get(cid, {}).get(
        "name", cid
    )
    for cid in sorted(by_company, key=name_of):
        for p in sorted(
            by_company[cid],
            key=lambda r: _CONF_RANK.get(r["confidence"], 0),
            reverse=True,
        ):
            event_name = data["event_by_id"].get(
                p["event_id"], {}
            ).get("name", p["event_id"])
            rows.append([
                name_of(cid),
                event_name,
                p.get("role", ""),
                p.get("confidence", ""),
                p.get("extracted_by", ""),
                p.get("evidence_url", ""),
                (p.get("evidence_excerpt") or "")[:300],
                p.get("extracted_at", ""),
            ])
    _write_sheet(
        wb, "Raw evidence",
        [
            "Company", "Event", "Role", "Confidence",
            "Extracted by", "Evidence URL", "Evidence excerpt",
            "Extracted at",
        ],
        rows,
        widths=[28, 38, 14, 16, 28, 50, 60, 18],
    )
    return len(rows)


# ---- Main ----

def build(
    path: Path | None = None,
    *,
    target_participants_only: bool = True,
    exclude_primes: bool = True,
) -> Path:
    """Write the multi-sheet XLSX. Returns the output path."""
    raw = _load_all()
    data = (
        _filter_data_for_target_audience(raw, exclude_primes=exclude_primes)
        if target_participants_only else raw
    )

    wb = Workbook()
    # openpyxl gives us an empty default sheet — drop it.
    wb.remove(wb.active)

    counts = {
        "Outreach":     _sheet_outreach(wb, data),
        "Master list":  _sheet_master_list(wb, data),
        "Cross-event":  _sheet_cross_event(wb, data),
        "Emerging":     _sheet_emerging(wb, data),
        "Raw evidence": _sheet_raw_evidence(wb, data),
    }

    out_dir = Path(__file__).parent / "out"
    out_dir.mkdir(parents=True, exist_ok=True)
    path = path or (
        out_dir / f"report_{date.today().isoformat()}.xlsx"
    )
    wb.save(path)
    return path


def main() -> None:
    path = build()
    print(f"wrote {path}")


if __name__ == "__main__":
    main()
